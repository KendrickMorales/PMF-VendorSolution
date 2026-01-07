import os
import tempfile
import shutil
import zipfile
import csv
import io
import re
from datetime import datetime
from pathlib import Path
from flask import request, jsonify, send_from_directory, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from collections import Counter

from mappings import load_mappings, save_mappings, check_file_mapping_status
from filename_generator import (
    is_part_number_filename,
    find_original_filename_by_part_number,
    get_or_create_part_mapping
)
from solidworks import (
    read_solidworks_properties,
    update_solidworks_property,
    SOLIDWORKS_AVAILABLE
)

def register_routes(app):
    """Register all routes with the Flask app"""
    
    @app.route('/api/scan-folder', methods=['POST'])
    def scan_folder():
        """Scan folder for SOLIDWORKS files and check for existing mappings"""
        data = request.json
        folder_path = data.get('folderPath')
        include_subdirectories = data.get('includeSubdirectories', False)  # Default to False
        
        if not folder_path or not os.path.exists(folder_path):
            return jsonify({'error': 'Invalid folder path'}), 400
        
        # Load existing mappings
        mappings = load_mappings()
        
        # Find all SOLIDWORKS files and related formats
        solidworks_extensions = ['.sldprt', '.sldasm', '.slddrw', '.step', '.stp', '.x_t', '.x_b']
        files = []
        processed_count = 0
        new_count = 0
        
        has_renamed_files = False
        for ext in solidworks_extensions:
            # Use recursive pattern if subdirectories should be included
            if include_subdirectories:
                pattern = f'**/*{ext}'
            else:
                pattern = f'*{ext}'
            
            for file_path in Path(folder_path).glob(pattern):
                file_str = str(file_path)
                # For subdirectories, preserve relative path; otherwise just filename
                if include_subdirectories:
                    relative_path = os.path.relpath(file_str, folder_path)
                else:
                    relative_path = file_path.name  # Just the filename, no subdirectory path
                file_name = file_path.name
                
                # Check if filename is a part number (12 digits)
                is_renamed_file = is_part_number_filename(file_name)
                original_filenames = []
                
                if is_renamed_file:
                    has_renamed_files = True
                    # Extract part number from filename
                    part_number = os.path.splitext(file_name)[0]
                    # Find original filename(s) with this part number
                    original_filenames = find_original_filename_by_part_number(part_number, mappings)
                    # Set mapping status based on found original files
                    if original_filenames:
                        mapping_status = {
                            'hasMapping': True,
                            'vendorPartNumber': part_number,
                            'basePartNumber': part_number[:9],
                            'revision': int(part_number[9:])
                        }
                        processed_count += 1
                    else:
                        mapping_status = {'hasMapping': False}
                        new_count += 1
                else:
                    # Check for existing mapping (try both absolute and relative paths)
                    mapping_status = check_file_mapping_status(file_str, mappings)
                    if not mapping_status['hasMapping']:
                        mapping_status = check_file_mapping_status(relative_path, mappings)
                    
                    if mapping_status['hasMapping']:
                        processed_count += 1
                    else:
                        new_count += 1
                
                files.append({
                    'name': file_name,
                    'path': file_str,
                    'relativePath': relative_path,
                    'hasMapping': mapping_status['hasMapping'],
                    'existingPartNumber': mapping_status.get('vendorPartNumber'),
                    'existingBase': mapping_status.get('basePartNumber'),
                    'existingRevision': mapping_status.get('revision'),
                    'isRenamedFile': is_renamed_file,
                    'originalFilenames': original_filenames
                })
        
        return jsonify({
            'files': files,
            'folderStatus': {
                'totalFiles': len(files),
                'processedFiles': processed_count,
                'newFiles': new_count,
                'hasBeenProcessed': processed_count > 0,
                'hasRenamedFiles': has_renamed_files
            }
        })

    @app.route('/api/process-uploaded-files', methods=['POST'])
    def process_uploaded_files():
        """Process uploaded SOLIDWORKS files"""
        if 'files' not in request.files:
            return jsonify({'error': 'No files uploaded'}), 400
        
        uploaded_files = request.files.getlist('files')
        if not uploaded_files or uploaded_files[0].filename == '':
            return jsonify({'error': 'No files selected'}), 400
        
        # Create temporary directory
        temp_dir = tempfile.mkdtemp()
        results = []
        mappings = load_mappings()
        
        try:
            # Save uploaded files temporarily
            file_paths = []
            for file in uploaded_files:
                if file.filename:
                    # Extract relative path from filename (includes folder structure)
                    relative_path = file.filename
                    temp_path = os.path.join(temp_dir, relative_path)
                    
                    # Create directory structure if needed
                    os.makedirs(os.path.dirname(temp_path), exist_ok=True)
                    
                    # Save file
                    file.save(temp_path)
                    file_paths.append({
                        'temp_path': temp_path,
                        'relative_path': relative_path,
                        'name': os.path.basename(relative_path)
                    })
            
            # Process each file
            for file_info in file_paths:
                file_path = file_info['temp_path']
                file_name = file_info['name']
                relative_path = file_info['relative_path']
                
                # Check if filename is a part number (12 digits)
                is_renamed_file = is_part_number_filename(file_name)
                original_filenames = []
                
                if is_renamed_file:
                    # Extract part number from filename
                    part_number = os.path.splitext(file_name)[0]
                    # Find original filename(s) with this part number
                    original_filenames = find_original_filename_by_part_number(part_number, mappings)
                    # Set mapping status based on found original files
                    if original_filenames:
                        mapping_status = {
                            'hasMapping': True,
                            'vendorPartNumber': part_number,
                            'basePartNumber': part_number[:9],
                            'revision': int(part_number[9:])
                        }
                    else:
                        mapping_status = {'hasMapping': False}
                else:
                    # Check for existing mapping
                    mapping_status = check_file_mapping_status(relative_path, mappings)
                
                file_data = {
                    'path': relative_path,
                    'name': file_name,
                    'properties': {},
                    'vendorPartNumber': None,
                    'basePartNumber': None,
                    'revision': 1,
                    'hasMapping': mapping_status['hasMapping'],
                    'isNew': not mapping_status['hasMapping'],
                    'isRenamedFile': is_renamed_file,
                    'originalFilenames': original_filenames
                }
                
                # Read existing properties
                properties = read_solidworks_properties(file_path)
                if properties:
                    file_data['properties'] = properties
                    if 'Vendor Part Number' in properties:
                        existing_part_number = properties['Vendor Part Number'].get('resolved', '')
                        if existing_part_number and len(existing_part_number) == 12:
                            file_data['vendorPartNumber'] = existing_part_number
                            file_data['basePartNumber'] = existing_part_number[:9]
                            file_data['revision'] = int(existing_part_number[9:])
                else:
                    # If SOLIDWORKS API not available, use file metadata
                    stat = os.stat(file_path)
                    modified_date = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    file_size_mb = stat.st_size / (1024 * 1024)
                    file_data['properties'] = {
                        'File Size': {'value': str(stat.st_size), 'resolved': f"{file_size_mb:.2f} MB"},
                        'Modified Date': {'value': str(stat.st_mtime), 'resolved': modified_date}
                    }
                    if not SOLIDWORKS_AVAILABLE:
                        file_data['note'] = 'SOLIDWORKS API not available - using file metadata only'
                
                # Use existing mapping if available, otherwise get or create
                if mapping_status['hasMapping']:
                    file_data['vendorPartNumber'] = mapping_status['vendorPartNumber']
                    file_data['basePartNumber'] = mapping_status['basePartNumber']
                    file_data['revision'] = mapping_status['revision']
                elif not file_data['vendorPartNumber']:
                    # Get or create part mapping (only if not already set from properties)
                    mapping, full_part_number = get_or_create_part_mapping(relative_path, mappings)
                    file_data['vendorPartNumber'] = full_part_number
                    file_data['basePartNumber'] = mapping['base']
                    file_data['revision'] = mapping['revision']
                else:
                    # Update mapping with existing part number from properties
                    base = file_data['basePartNumber']
                    rev = file_data['revision']
                    mappings[relative_path] = {'base': base, 'revision': rev}
                
                save_mappings(mappings)
                
                results.append(file_data)
            
            return jsonify({'results': results})
        
        finally:
            # Clean up temporary directory
            try:
                shutil.rmtree(temp_dir)
            except Exception as e:
                print(f"Error cleaning up temp directory: {e}")

    @app.route('/api/read-properties', methods=['POST'])
    def read_properties():
        """Read properties from SOLIDWORKS files"""
        data = request.json
        file_paths = data.get('filePaths', [])
        
        results = []
        mappings = load_mappings()
        
        for file_path in file_paths:
            file_name = os.path.basename(file_path)
            
            # Check if filename is a part number (12 digits)
            is_renamed_file = is_part_number_filename(file_name)
            original_filenames = []
            
            if is_renamed_file:
                # Extract part number from filename
                part_number = os.path.splitext(file_name)[0]
                # Find original filename(s) with this part number
                original_filenames = find_original_filename_by_part_number(part_number, mappings)
                # Set mapping status based on found original files
                if original_filenames:
                    mapping_status = {
                        'hasMapping': True,
                        'vendorPartNumber': part_number,
                        'basePartNumber': part_number[:9],
                        'revision': int(part_number[9:])
                    }
                else:
                    mapping_status = {'hasMapping': False}
            else:
                # Check for existing mapping
                mapping_status = check_file_mapping_status(file_path, mappings)
            
            file_info = {
                'path': file_path,
                'name': file_name,
                'properties': {},
                'vendorPartNumber': None,
                'basePartNumber': None,
                'revision': 1,
                'hasMapping': mapping_status['hasMapping'],
                'isNew': not mapping_status['hasMapping'],
                'isRenamedFile': is_renamed_file,
                'originalFilenames': original_filenames
            }
            
            # Read existing properties
            properties = read_solidworks_properties(file_path)
            if properties:
                file_info['properties'] = properties
                # Check if vendor part number already exists
                if 'Vendor Part Number' in properties:
                    existing_part_number = properties['Vendor Part Number'].get('resolved', '')
                    if existing_part_number and len(existing_part_number) == 12:
                        file_info['vendorPartNumber'] = existing_part_number
                        file_info['basePartNumber'] = existing_part_number[:9]
                        file_info['revision'] = int(existing_part_number[9:])
            else:
                # If SOLIDWORKS API not available, use file metadata
                stat = os.stat(file_path)
                modified_date = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                file_size_mb = stat.st_size / (1024 * 1024)
                file_info['properties'] = {
                    'File Size': {'value': str(stat.st_size), 'resolved': f"{file_size_mb:.2f} MB"},
                    'Modified Date': {'value': str(stat.st_mtime), 'resolved': modified_date}
                }
                if not SOLIDWORKS_AVAILABLE:
                    file_info['note'] = 'SOLIDWORKS API not available - using file metadata only'
            
            # Use existing mapping if available, otherwise get or create
            if mapping_status['hasMapping']:
                file_info['vendorPartNumber'] = mapping_status['vendorPartNumber']
                file_info['basePartNumber'] = mapping_status['basePartNumber']
                file_info['revision'] = mapping_status['revision']
            elif not file_info['vendorPartNumber']:
                # Get or create part mapping (only if not already set from properties)
                mapping, full_part_number = get_or_create_part_mapping(file_path, mappings)
                file_info['vendorPartNumber'] = full_part_number
                file_info['basePartNumber'] = mapping['base']
                file_info['revision'] = mapping['revision']
            else:
                # Update mapping with existing part number from properties
                base = file_info['basePartNumber']
                rev = file_info['revision']
                mappings[file_path] = {'base': base, 'revision': rev}
            
            save_mappings(mappings)
            
            results.append(file_info)
        
        return jsonify({'results': results})

    @app.route('/api/create-revision', methods=['POST'])
    def create_revision():
        """Create a new revision for an existing part"""
        data = request.json
        file_path = data.get('filePath')
        new_revision = data.get('revision')  # Optional, will auto-increment if not provided
        
        if not file_path:
            return jsonify({'error': 'File path required'}), 400
        
        mappings = load_mappings()
        
        if file_path not in mappings:
            return jsonify({'error': 'File not found in mappings'}), 404
        
        # Get current mapping
        current_mapping = mappings[file_path]
        
        # Handle old format migration
        if isinstance(current_mapping, str):
            if len(current_mapping) == 12:
                base = current_mapping[:9]
                rev = int(current_mapping[9:])
                current_mapping = {'base': base, 'revision': rev}
            else:
                return jsonify({'error': 'Invalid mapping format'}), 400
        
        # Determine new revision number
        if new_revision is None:
            new_revision = current_mapping.get('revision', 1) + 1
        
        # Update mapping with new revision
        current_mapping['revision'] = new_revision
        mappings[file_path] = current_mapping
        save_mappings(mappings)
        
        full_part_number = f"{current_mapping['base']}{new_revision:03d}"
        
        return jsonify({
            'basePartNumber': current_mapping['base'],
            'revision': new_revision,
            'vendorPartNumber': full_part_number
        })

    @app.route('/api/update-properties', methods=['POST'])
    def update_properties():
        """Update vendor part numbers in SOLIDWORKS files"""
        data = request.json
        updates = data.get('updates', [])  # List of {filePath, vendorPartNumber, revision}
        
        mappings = load_mappings()
        results = []
        
        for update in updates:
            file_path = update['filePath']
            vendor_part_number = update['vendorPartNumber']
            revision = update.get('revision')
            
            # Update mapping if revision is provided
            if revision is not None:
                mapping, _ = get_or_create_part_mapping(file_path, mappings, revision)
                vendor_part_number = f"{mapping['base']}{mapping['revision']:03d}"
            
            success = update_solidworks_property(file_path, 'Vendor Part Number', vendor_part_number)
            
            if success:
                results.append({
                    'filePath': file_path,
                    'success': True
                })
            else:
                # Even if SOLIDWORKS update fails, save the mapping
                results.append({
                    'filePath': file_path,
                    'success': not SOLIDWORKS_AVAILABLE,  # Success if API not available (mapping saved)
                    'error': 'Failed to update property in file' if SOLIDWORKS_AVAILABLE else 'Mapping saved (SOLIDWORKS API not available)',
                    'note': 'Use SOLIDWORKS to manually update files or install SOLIDWORKS Document Manager API' if not SOLIDWORKS_AVAILABLE else None
                })
        
        save_mappings(mappings)
        return jsonify({'results': results})

    @app.route('/api/mappings', methods=['GET'])
    def get_mappings():
        """Get all vendor part number mappings"""
        return jsonify(load_mappings())

    @app.route('/api/export-mappings', methods=['GET'])
    def export_mappings():
        """Export mappings as CSV"""
        mappings = load_mappings()
        csv_lines = ['File Path,Base Part Number,Revision,Vendor Part Number']
        for file_path, mapping_data in mappings.items():
            if isinstance(mapping_data, dict):
                base = mapping_data.get('base', '')
                rev = mapping_data.get('revision', 1)
                full = f"{base}{rev:03d}"
                csv_lines.append(f'"{file_path}","{base}","{rev}","{full}"')
            else:
                # Old format - migrate
                if len(str(mapping_data)) == 12:
                    base = str(mapping_data)[:9]
                    rev = int(str(mapping_data)[9:])
                    csv_lines.append(f'"{file_path}","{base}","{rev}","{mapping_data}"')
                else:
                    csv_lines.append(f'"{file_path}","","","{mapping_data}"')
        return jsonify({'csv': '\n'.join(csv_lines)})

    @app.route('/api/generate-files-with-part-numbers', methods=['POST'])
    def generate_files_with_part_numbers():
        """Generate folder with renamed files and Excel file"""
        data = request.json
        folder_path = data.get('folderPath')
        files_data = data.get('files', [])  # List of {path, name, vendorPartNumber}
        
        if not folder_path or not files_data:
            return jsonify({'error': 'Folder path and files data required'}), 400
        
        try:
            # Create output folder
            output_folder = os.path.join(folder_path, 'Vendor_Part_Numbers')
            os.makedirs(output_folder, exist_ok=True)
            
            # Count part numbers for quantity
            part_number_counts = Counter()
            file_extensions = {}
            
            # Copy files with part number as name
            copied_files = []
            for file_info in files_data:
                original_path = file_info.get('path')
                vendor_part_number = file_info.get('vendorPartNumber')
                
                if not vendor_part_number or not original_path:
                    continue
                
                # Get original file extension
                original_ext = os.path.splitext(original_path)[1] or os.path.splitext(file_info.get('name', ''))[1]
                if not original_ext:
                    # Try to determine from file_info name
                    original_ext = os.path.splitext(file_info.get('name', ''))[1]
                
                # Build full original path
                if os.path.isabs(original_path):
                    full_original_path = original_path
                else:
                    full_original_path = os.path.join(folder_path, original_path)
                
                # Check if file exists
                if not os.path.exists(full_original_path):
                    # Try with just filename
                    full_original_path = os.path.join(folder_path, os.path.basename(original_path))
                    if not os.path.exists(full_original_path):
                        continue
                
                # New filename: part number + original extension
                new_filename = f"{vendor_part_number}{original_ext}"
                new_file_path = os.path.join(output_folder, new_filename)
                
                # Copy file
                try:
                    shutil.copy2(full_original_path, new_file_path)
                    copied_files.append(new_filename)
                    part_number_counts[vendor_part_number] += 1
                    file_extensions[vendor_part_number] = original_ext
                except Exception as e:
                    print(f"Error copying file {full_original_path}: {e}")
                    continue
            
            # Generate Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Part Numbers"
            
            # Headers
            ws['A1'] = 'ITEM NO.'
            ws['B1'] = 'PART NUMBER'
            ws['C1'] = 'QTY'
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font_color = Font(bold=True, color="FFFFFF")
            
            ws['A1'].font = header_font_color
            ws['A1'].fill = header_fill
            ws['B1'].font = header_font_color
            ws['B1'].fill = header_fill
            ws['C1'].font = header_font_color
            ws['C1'].fill = header_fill
            
            # Add data (unique part numbers only, sorted)
            row = 2
            item_no = 1
            for part_number in sorted(set(part_number_counts.keys())):
                ws[f'A{row}'] = item_no
                ws[f'B{row}'] = part_number
                ws[f'C{row}'] = ''  # QTY left blank for manual entry
                row += 1
                item_no += 1
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            
            # Save Excel file
            excel_path = os.path.join(output_folder, 'Part_Numbers_List.xlsx')
            wb.save(excel_path)
            
            return jsonify({
                'success': True,
                'outputFolder': output_folder,
                'filesCopied': len(copied_files),
                'uniquePartNumbers': len(set(part_number_counts.keys())),
                'excelFile': excel_path
            })
        
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/generate-files-from-upload', methods=['POST'])
    def generate_files_from_upload():
        """Generate folder with renamed files and Excel file from uploaded files"""
        if 'files' not in request.files:
            return jsonify({'error': 'No files uploaded'}), 400
        
        uploaded_files = request.files.getlist('files')
        
        if not uploaded_files:
            return jsonify({'error': 'No files selected'}), 400
        
        try:
            # Create temporary directory for processing
            temp_dir = tempfile.mkdtemp()
            output_folder = os.path.join(temp_dir, 'Vendor_Part_Numbers')
            os.makedirs(output_folder, exist_ok=True)
            
            # Count part numbers for quantity
            part_number_counts = Counter()
            
            # Save uploaded files with part number names
            copied_files = []
            for file in uploaded_files:
                if not file.filename:
                    continue
                
                # Extract part number from filename (format: partnumber.ext)
                # The filename is already set to partnumber.ext in FormData
                vendor_part_number = os.path.splitext(file.filename)[0]
                
                # New filename is already set in FormData
                new_filename = file.filename
                new_file_path = os.path.join(output_folder, new_filename)
                
                # Save file
                try:
                    file.save(new_file_path)
                    copied_files.append(new_filename)
                    part_number_counts[vendor_part_number] += 1
                except Exception as e:
                    print(f"Error saving file {new_filename}: {e}")
                    continue
            
            # Generate Excel file
            wb = Workbook()
            ws = wb.active
            ws.title = "Part Numbers"
            
            # Headers
            ws['A1'] = 'ITEM NO.'
            ws['B1'] = 'PART NUMBER'
            ws['C1'] = 'QTY'
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font_color = Font(bold=True, color="FFFFFF")
            
            ws['A1'].font = header_font_color
            ws['A1'].fill = header_fill
            ws['B1'].font = header_font_color
            ws['B1'].fill = header_fill
            ws['C1'].font = header_font_color
            ws['C1'].fill = header_fill
            
            # Add data (unique part numbers only, sorted)
            row = 2
            item_no = 1
            for part_number in sorted(set(part_number_counts.keys())):
                ws[f'A{row}'] = item_no
                ws[f'B{row}'] = part_number
                ws[f'C{row}'] = ''  # QTY left blank for manual entry
                row += 1
                item_no += 1
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            
            # Save Excel file
            excel_path = os.path.join(output_folder, 'Part_Numbers_List.xlsx')
            wb.save(excel_path)
            
            # Create zip file with all generated files
            zip_path = os.path.join(temp_dir, 'Vendor_Part_Numbers.zip')
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(output_folder):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, output_folder)
                        zipf.write(file_path, arcname)
            
            # Read zip file and return it
            with open(zip_path, 'rb') as f:
                zip_data = f.read()
            
            # Clean up temp directory
            shutil.rmtree(temp_dir, ignore_errors=True)
            
            response = Response(
                zip_data,
                mimetype='application/zip',
                headers={
                    'Content-Disposition': 'attachment; filename=Vendor_Part_Numbers.zip'
                }
            )
            return response
        
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/api/all-processed-files', methods=['GET'])
    def get_all_processed_files():
        """Get all processed files from mappings"""
        mappings = load_mappings()
        
        processed_files = []
        for file_path, mapping_data in mappings.items():
            if isinstance(mapping_data, dict):
                base = mapping_data.get('base', '')
                rev = mapping_data.get('revision', 1)
                full_part_number = f"{base}{rev:03d}"
                
                processed_files.append({
                    'originalPath': file_path,
                    'originalName': os.path.basename(file_path),
                    'basePartNumber': base,
                    'revision': rev,
                    'fullPartNumber': full_part_number
                })
            elif isinstance(mapping_data, str) and len(mapping_data) == 12:
                # Handle old format
                processed_files.append({
                    'originalPath': file_path,
                    'originalName': os.path.basename(file_path),
                    'basePartNumber': mapping_data[:9],
                    'revision': int(mapping_data[9:]),
                    'fullPartNumber': mapping_data
                })
        
        return jsonify({'files': processed_files, 'total': len(processed_files)})

    @app.route('/api/match-csv', methods=['POST'])
    def match_csv():
        """Match CSV entries against processed files"""
        if 'csvFile' not in request.files:
            return jsonify({'error': 'No CSV file uploaded'}), 400
        
        csv_file = request.files['csvFile']
        if csv_file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        try:
            # Read CSV file
            stream = io.StringIO(csv_file.stream.read().decode("UTF8"), newline=None)
            csv_reader = csv.DictReader(stream)
            
            # Get all processed files
            mappings = load_mappings()
            processed_files = []
            for file_path, mapping_data in mappings.items():
                if isinstance(mapping_data, dict):
                    base = mapping_data.get('base', '')
                    rev = mapping_data.get('revision', 1)
                    full_part_number = f"{base}{rev:03d}"
                    processed_files.append({
                        'originalPath': file_path,
                        'originalName': os.path.basename(file_path),
                        'basePartNumber': base,
                        'revision': rev,
                        'fullPartNumber': full_part_number
                    })
                elif isinstance(mapping_data, str) and len(mapping_data) == 12:
                    processed_files.append({
                        'originalPath': file_path,
                        'originalName': os.path.basename(file_path),
                        'basePartNumber': mapping_data[:9],
                        'revision': int(mapping_data[9:]),
                        'fullPartNumber': mapping_data
                    })
            
            # Create lookup dictionaries
            by_full_part = {f['fullPartNumber']: f for f in processed_files}
            by_base_part = {f['basePartNumber']: f for f in processed_files}
            by_filename = {f['originalName'].lower(): f for f in processed_files}
            
            # Also create lookup by filename without extension
            by_filename_no_ext = {}
            for f in processed_files:
                name_no_ext = os.path.splitext(f['originalName'])[0].lower()
                if name_no_ext not in by_filename_no_ext:
                    by_filename_no_ext[name_no_ext] = f
            
            # Match CSV entries
            matches = []
            unmatched = []
            csv_rows = list(csv_reader)
            
            for row in csv_rows:
                # Try to match by various fields
                matched = None
                match_type = None
                
                # Try full part number (12 digits) - exact match
                for key, value in row.items():
                    if value:
                        value_str = str(value).strip()
                        # Check if entire value is a 12-digit number
                        if len(value_str) == 12 and value_str.isdigit():
                            if value_str in by_full_part:
                                matched = by_full_part[value_str]
                                match_type = 'fullPartNumber'
                                break
                        # Check if value contains a 12-digit number
                        part_num_match = re.search(r'\d{12}', value_str)
                        if part_num_match:
                            part_num = part_num_match.group()
                            if part_num in by_full_part:
                                matched = by_full_part[part_num]
                                match_type = 'fullPartNumber'
                                break
                
                # Try base part number (9 digits) - exact or embedded
                if not matched:
                    for key, value in row.items():
                        if value:
                            value_str = str(value).strip()
                            # Check if entire value is a 9-digit number
                            if len(value_str) == 9 and value_str.isdigit():
                                if value_str in by_base_part:
                                    matched = by_base_part[value_str]
                                    match_type = 'basePartNumber'
                                    break
                            # Check if value contains a 9-digit number
                            part_num_match = re.search(r'\d{9}', value_str)
                            if part_num_match:
                                part_num = part_num_match.group()
                                if part_num in by_base_part:
                                    matched = by_base_part[part_num]
                                    match_type = 'basePartNumber'
                                    break
                
                # Try exact filename match (with extension)
                if not matched:
                    for key, value in row.items():
                        if value:
                            filename = str(value).strip().lower()
                            if filename in by_filename:
                                matched = by_filename[filename]
                                match_type = 'filename'
                                break
                
                # Try filename without extension match
                if not matched:
                    for key, value in row.items():
                        if value:
                            value_str = str(value).strip().lower()
                            # Remove extension if present
                            value_no_ext = os.path.splitext(value_str)[0]
                            if value_no_ext in by_filename_no_ext:
                                matched = by_filename_no_ext[value_no_ext]
                                match_type = 'filename'
                                break
                
                # Try extracting filename from CSV (filename is usually at the start before first space)
                # This handles cases like "SHRD-FSCM-PLT-MS-101-LH DESCRIPTION TEXT"
                if not matched:
                    for key, value in row.items():
                        if value:
                            value_str = str(value).strip()
                            # Extract potential filename - take first "word" (before first space)
                            # or match pattern like "SHRD-FSCM-PLT-MS-101-LH"
                            parts = value_str.split()
                            if parts:
                                potential_filename = parts[0].lower()
                                potential_filename_no_ext = os.path.splitext(potential_filename)[0]
                                
                                # Try exact match with extension
                                if potential_filename in by_filename:
                                    matched = by_filename[potential_filename]
                                    match_type = 'filename'
                                    break
                                
                                # Try match without extension
                                if potential_filename_no_ext in by_filename_no_ext:
                                    matched = by_filename_no_ext[potential_filename_no_ext]
                                    match_type = 'filename'
                                    break
                                
                                # Try matching against all processed filenames (check if processed filename starts with CSV filename)
                                for proc_file in processed_files:
                                    proc_filename = proc_file['originalName'].lower()
                                    proc_filename_no_ext = os.path.splitext(proc_filename)[0].lower()
                                    
                                    # Check if processed filename matches the extracted CSV filename
                                    if (potential_filename == proc_filename or 
                                        potential_filename_no_ext == proc_filename_no_ext or
                                        potential_filename == proc_filename_no_ext or
                                        potential_filename_no_ext == proc_filename):
                                        matched = proc_file
                                        match_type = 'filename'
                                        break
                                    
                                    # Check if processed filename starts with CSV filename part
                                    if (proc_filename.startswith(potential_filename) or
                                        proc_filename_no_ext.startswith(potential_filename_no_ext)):
                                        matched = proc_file
                                        match_type = 'filename'
                                        break
                                    
                                    # Check if CSV filename part is in processed filename
                                    if (potential_filename in proc_filename or
                                        potential_filename_no_ext in proc_filename_no_ext):
                                        matched = proc_file
                                        match_type = 'filename'
                                        break
                                
                                if matched:
                                    break
                
                # Try reverse: check if CSV value starts with any processed filename
                if not matched:
                    for key, value in row.items():
                        if value:
                            value_str = str(value).strip().lower()
                            value_no_ext = os.path.splitext(value_str)[0]
                            
                            for proc_file in processed_files:
                                proc_filename = proc_file['originalName'].lower()
                                proc_filename_no_ext = os.path.splitext(proc_filename)[0].lower()
                                
                                # Check if CSV value starts with processed filename
                                if (value_str.startswith(proc_filename) or 
                                    value_str.startswith(proc_filename_no_ext) or
                                    value_no_ext.startswith(proc_filename) or
                                    value_no_ext.startswith(proc_filename_no_ext)):
                                    matched = proc_file
                                    match_type = 'filename'
                                    break
                            
                            if matched:
                                break
                
                if matched:
                    matches.append({
                        'csvRow': row,
                        'matchedFile': matched,
                        'matchType': match_type
                    })
                else:
                    unmatched.append(row)
            
            return jsonify({
                'matches': matches,
                'unmatched': unmatched,
                'totalCsvRows': len(csv_rows),
                'matchedCount': len(matches),
                'unmatchedCount': len(unmatched)
            })
        
        except Exception as e:
            return jsonify({'error': f'Error processing CSV: {str(e)}'}), 500

    @app.route('/api/generate-excel-from-matches', methods=['POST'])
    def generate_excel_from_matches():
        """Generate Excel file from matched CSV entries"""
        data = request.json
        matches = data.get('matches', [])
        
        if not matches:
            return jsonify({'error': 'No matches provided'}), 400
        
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Matched Part Numbers"
            
            # Headers
            ws['A1'] = 'ITEM NO.'
            ws['B1'] = 'PART NUMBER'
            ws['C1'] = 'QTY'
            ws['D1'] = 'ORIGINAL FILENAME'
            ws['E1'] = 'ORIGINAL PATH'
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font_color = Font(bold=True, color="FFFFFF")
            
            for col in ['A1', 'B1', 'C1', 'D1', 'E1']:
                ws[col].font = header_font_color
                ws[col].fill = header_fill
            
            # Add data
            row = 2
            item_no = 1
            seen_part_numbers = set()
            
            for match in matches:
                matched_file = match['matchedFile']
                part_number = matched_file['fullPartNumber']
                
                # Only add unique part numbers
                if part_number not in seen_part_numbers:
                    ws[f'A{row}'] = item_no
                    ws[f'B{row}'] = part_number
                    ws[f'C{row}'] = ''  # QTY left blank for manual entry
                    ws[f'D{row}'] = matched_file['originalName']
                    ws[f'E{row}'] = matched_file['originalPath']
                    
                    seen_part_numbers.add(part_number)
                    row += 1
                    item_no += 1
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 50
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = Response(
                output.read(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': 'attachment; filename=Matched_Part_Numbers.xlsx'
                }
            )
            return response
        
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @app.route('/', defaults={'path': ''})
    @app.route('/<path:path>')
    def serve(path):
        """Serve React app"""
        if path != "" and os.path.exists(os.path.join(app.static_folder, path)):
            return send_from_directory(app.static_folder, path)
        else:
            return send_from_directory(app.static_folder, 'index.html')

